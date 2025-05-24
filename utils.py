import win32com.client
from openpyxl import Workbook, load_workbook
import os
import sys
import re
from datetime import datetime, timedelta


class OutlookHandler:
    '''Класс для работы с Outlook.'''

    def __init__(self):
        self.outlook = win32com.client.Dispatch('Outlook.Application').GetNamespace('MAPI')

    def get_mailbox(self, mailbox_name):
        try:
            return self.outlook.Folders[mailbox_name]
        except Exception as e:
            raise ValueError(f'Почтовый ящик \'{mailbox_name}\' не найден: {e}')

    def find_folder(self, parent_folder, target_name):
        if parent_folder.Name == target_name:
            return parent_folder
        for folder in parent_folder.Folders:
            found = self.find_folder(folder, target_name)
            if found:
                return found
        return None


class ExcelHandler:
    '''Класс для работы с Excel.'''

    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.workbook = None
        self.worksheet = None
        self.load_or_create_workbook()

    def load_or_create_workbook(self):
        if os.path.exists(self.excel_path):
            self.workbook = load_workbook(self.excel_path)
            self.worksheet = self.workbook.active
        else:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = 'Темы писем'
            self.worksheet.append(['№', 'Уровень угрозы', 'INC-номер', 'Тема', 'Дата регистрации',
                                   'Время инцидента', 'Дата получения письма'])

    def add_data_to_excel(self, data):
        self.worksheet.insert_rows(2)
        for col_num, cell_value in enumerate(data, start=1):
            self.worksheet.cell(row=2, column=col_num, value=cell_value)

    def save_workbook(self):
        self.workbook.save(self.excel_path)


def parse_email_data(message):
    '''Извлекает и форматирует данные из письма.'''
    subject = message.Subject
    body = message.Body

    match_subject = re.match(r'(\w+)\s+([A-Z]+-\d+)\s+\'(.*)\'', subject)
    if match_subject:
        word1, inc_number, quoted_text = match_subject.groups()
    else:
        word1, inc_number, quoted_text = 'N/A', 'N/A', 'N/A'

    match_date1 = re.search(r'Дата регистрации\s*(\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}:\d{2})', body)
    date_reg1 = 'N/A'
    if match_date1:
        try:
            date_str = match_date1.group(1)
            date_obj = datetime.strptime(date_str, '%d.%m.%Y %H:%M:%S')
            date_obj += timedelta(hours=3)
            date_reg1 = date_obj.strftime('%d.%m.%Y %H:%M:%S')
        except ValueError:
            date_reg1 = 'Некорректная дата'

    match_date2 = re.search(
        r'(\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4},\s+\d{2}:\d{2}:\d{2})\s+UTC', body)
    date_reg2 = 'N/A'
    if match_date2:
        try:
            date_obj = datetime.strptime(match_date2.group(1), '%d %b %Y, %H:%M:%S')
            date_obj += timedelta(hours=3)
            date_reg2 = date_obj.strftime('%d.%m.%Y %H:%M:%S')
        except ValueError:
            date_reg2 = 'Некорректная дата'

    date_received = message.ReceivedTime.strftime('%d.%m.%Y %H:%M:%S') if hasattr(
        message, 'ReceivedTime') else 'N/A'

    return word1, inc_number, quoted_text, date_reg1, date_reg2, date_received


def get_parameters_from_file(config_file='config.txt'):
    params = {}
    try:
        if getattr(sys, 'frozen', False):  # Если запущен как .exe
            base_dir = os.path.dirname(sys.executable)
        else:  # Если запущен как .py
            base_dir = os.path.dirname(os.path.abspath(__file__))
        
        config_path = os.path.join(base_dir, config_file)
        
        with open(config_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if '=' in line:
                    key, value = line.split('=', 1)
                    params[key.strip()] = value.strip()
    except FileNotFoundError:
        print(f'Файл конфигурации \'{config_path}\' не найден. Положите config.txt в папку: {base_dir}')
        return None
    except Exception as e:
        print(f'Ошибка при чтении файла конфигурации: {e}')
        return None
    return params
